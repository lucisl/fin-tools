import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill

SHOPPING_LIST_PATH = "config/shopping_list.json"
CATEGORIES_PATH = "config/categories.json"


def load_data():
    """加载购物清单数据"""
    if not os.path.exists(SHOPPING_LIST_PATH):
        return []
    with open(SHOPPING_LIST_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_data(data):
    """保存购物清单数据"""
    os.makedirs("config", exist_ok=True)
    with open(SHOPPING_LIST_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def export_to_excel(items):
    """导出购物清单到Excel"""
    df = pd.DataFrame(items)
    
    if df.empty:
        return None
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="购物清单")
        
        workbook = writer.book
        sheet = writer.sheets["购物清单"]
        
        column_widths = {
            "A": 20,
            "B": 12,
            "C": 8,
            "D": 8,
            "E": 12,
            "F": 10,
            "G": 12,
            "H": 10,
            "I": 12,
            "J": 20,
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
    
    return output.getvalue()


def add_item(name, category, quantity, unit, budget_price, priority, platform, notes=""):
    """新增商品"""
    items = load_data()
    new_item = {
        "id": str(datetime.now().timestamp()),
        "name": name,
        "category": category,
        "quantity": quantity,
        "unit": unit,
        "budget_price": budget_price,
        "priority": priority,
        "platform": platform,
        "status": "待购买",
        "actual_price": 0,
        "notes": notes,
        "created_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    items.append(new_item)
    save_data(items)
    return new_item


def update_item(item_id, **kwargs):
    """更新商品信息"""
    items = load_data()
    for item in items:
        if item["id"] == item_id:
            item.update(kwargs)
            item["updated_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    save_data(items)


def delete_item(item_id):
    """删除商品"""
    items = load_data()
    items = [item for item in items if item["id"] != item_id]
    save_data(items)


def toggle_status(item_id):
    """切换购买状态"""
    items = load_data()
    status_cycle = ["待购买", "已购买", "已取消"]
    for item in items:
        if item["id"] == item_id:
            current_index = status_cycle.index(item["status"])
            next_index = (current_index + 1) % len(status_cycle)
            item["status"] = status_cycle[next_index]
            item["updated_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    save_data(items)


def get_items(category=None, priority=None, status=None):
    """获取商品列表,支持筛选"""
    items = load_data()
    
    if category and category != "全部":
        items = [item for item in items if item["category"] == category]
    
    if priority and priority != "全部":
        items = [item for item in items if item["priority"] == priority]
    
    if status and status != "全部":
        items = [item for item in items if item["status"] == status]
    
    return items


def load_categories():
    """加载分类列表"""
    if not os.path.exists(CATEGORIES_PATH):
        return []
    with open(CATEGORIES_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
        return data.get("categories", [])


def save_categories(categories):
    """保存分类列表"""
    os.makedirs("config", exist_ok=True)
    with open(CATEGORIES_PATH, "w", encoding="utf-8") as f:
        json.dump({"categories": categories}, f, ensure_ascii=False, indent=2)


def add_category(category_name):
    """新增分类"""
    categories = load_categories()
    if category_name not in categories:
        categories.append(category_name)
        save_categories(categories)
        return True
    return False


def delete_category(category_name):
    """删除分类"""
    items = load_data()
    has_items = any(item["category"] == category_name for item in items)
    if has_items:
        return False
    
    categories = load_categories()
    if category_name in categories:
        categories.remove(category_name)
        save_categories(categories)
        return True
    return False


def calculate_budget_summary(items):
    """计算预算汇总"""
    total_budget = sum(item.get("budget_price", 0) * item.get("quantity", 1) for item in items)
    spent = sum(item.get("actual_price", 0) for item in items if item["status"] == "已购买")
    remaining = total_budget - spent
    return {
        "total_budget": total_budget,
        "spent": spent,
        "remaining": remaining
    }


def calculate_category_stats(items):
    """按分类统计"""
    stats = {}
    for item in items:
        category = item.get("category", "未分类")
        if category not in stats:
            stats[category] = {
                "count": 0,
                "budget": 0,
                "spent": 0
            }
        stats[category]["count"] += 1
        stats[category]["budget"] += item.get("budget_price", 0) * item.get("quantity", 1)
        if item["status"] == "已购买":
            stats[category]["spent"] += item.get("actual_price", 0)
    return stats


def calculate_priority_stats(items):
    """按优先级统计"""
    stats = {
        "必须": {"count": 0, "budget": 0, "spent": 0},
        "想买": {"count": 0, "budget": 0, "spent": 0},
        "可选": {"count": 0, "budget": 0, "spent": 0}
    }
    
    for item in items:
        priority = item.get("priority", "可选")
        if priority in stats:
            stats[priority]["count"] += 1
            stats[priority]["budget"] += item.get("budget_price", 0) * item.get("quantity", 1)
            if item["status"] == "已购买":
                stats[priority]["spent"] += item.get("actual_price", 0)
    
    return stats


def calculate_completion_rate(items):
    """计算购买完成率"""
    if not items:
        return 0
    
    completed = sum(1 for item in items if item["status"] == "已购买")
    return (completed / len(items)) * 100


# ======================================
#               Web UI
# ======================================
st.title("🛒 618购物清单")

# 侧边栏筛选
st.sidebar.header("🔍 筛选条件")

# 加载分类
categories = load_categories()
category_filter = st.sidebar.selectbox("按分类筛选", ["全部"] + categories)

# 按优先级筛选
priority_filter = st.sidebar.selectbox("按优先级筛选", ["全部", "必须", "想买", "可选"])

# 按状态筛选
status_filter = st.sidebar.selectbox("按状态筛选", ["全部", "待购买", "已购买", "已取消"])

# 获取筛选后的商品
filtered_items = get_items(
    category=category_filter,
    priority=priority_filter,
    status=status_filter
)

# 侧边栏预算统计
st.sidebar.markdown("---")
st.sidebar.header("💰 预算统计")
budget_summary = calculate_budget_summary(load_data())
st.sidebar.metric("总预算", f"¥{budget_summary['total_budget']:.2f}")
st.sidebar.metric("已花费", f"¥{budget_summary['spent']:.2f}")
st.sidebar.metric("剩余预算", f"¥{budget_summary['remaining']:.2f}")

st.sidebar.markdown("---")
st.sidebar.header("📂 分类管理")

@st.dialog("新增分类")
def add_category_dialog():
    new_category = st.text_input("分类名称")
    
    col1, col2 = st.columns(2)
    if col1.button("保存", use_container_width=True):
        if not new_category.strip():
            st.error("分类名称不能为空")
        else:
            if add_category(new_category):
                st.success("新增成功!")
                st.rerun()
            else:
                st.warning("该分类已存在")
    
    if col2.button("取消", use_container_width=True):
        st.rerun()

if st.sidebar.button("➕ 新增分类", use_container_width=True):
    add_category_dialog()

categories = load_categories()
for cat in categories:
    cols = st.sidebar.columns([3, 1])
    cols[0].write(f"• {cat}")
    
    if cols[1].button("🗑️", key=f"del_cat_{cat}", help="删除分类"):
        if delete_category(cat):
            st.success(f"已删除分类: {cat}")
            st.rerun()
        else:
            st.error("该分类下有关联商品,无法删除")

tab1, tab2, tab3 = st.tabs(["📝 清单管理", "📊 统计分析", "📥 导出"])

with tab1:
    st.markdown("### 商品列表")
    
    @st.dialog("新增商品")
    def add_item_dialog():
        name = st.text_input("商品名称 *")
        categories = load_categories()
        category = st.selectbox("分类 *", categories)
        col1, col2 = st.columns(2)
        quantity = col1.number_input("数量 *", min_value=1, value=1)
        unit = col2.text_input("单位", value="个")
        budget_price = st.number_input("预算价格 *", min_value=0.0, step=0.01)
        priority = st.selectbox("优先级 *", ["必须", "想买", "可选"])
        platform = st.selectbox("购买平台", ["淘宝", "天猫", "京东", "拼多多", "其他"])
        notes = st.text_area("备注")
        
        col1, col2 = st.columns(2)
        if col1.button("保存", use_container_width=True):
            if not name.strip():
                st.error("商品名称不能为空")
            else:
                add_item(name, category, quantity, unit, budget_price, priority, platform, notes)
                st.success("新增成功!")
                st.rerun()
        
        if col2.button("取消", use_container_width=True):
            st.rerun()
    
    if st.button("➕ 新增商品", use_container_width=True):
        add_item_dialog()
    
    st.markdown("---")
    
    if not filtered_items:
        st.info("暂无商品,请点击上方按钮新增")
    else:
        for item in filtered_items:
            priority_emoji = {"必须": "🔴", "想买": "🟡", "可选": "⚪"}
            status_emoji = {"待购买": "🛒", "已购买": "✅", "已取消": "❌"}
            
            with st.container():
                cols = st.columns([4, 3, 3])
                
                with cols[0]:
                    st.write(f"**{item['name']}**")
                    st.caption(f"📦 {item['quantity']}{item['unit']} | 💰 ¥{item['budget_price']:.2f}")
                    st.caption(f"📂 {item['category']} | 🏪 {item['platform']}")
                
                with cols[1]:
                    st.write(f"{status_emoji.get(item['status'], '')} {item['status']}")
                    st.write(f"{priority_emoji.get(item['priority'], '')} {item['priority']}")
                
                with cols[2]:
                    c1, c2, c3 = st.columns(3)
                    
                    if c1.button("🔄", key=f"toggle_{item['id']}", help="切换状态"):
                        toggle_status(item['id'])
                        st.rerun()
                    
                    if c2.button("✏️", key=f"edit_{item['id']}", help="编辑"):
                        st.session_state.edit_item_id = item['id']
                        st.rerun()
                    
                    if c3.button("🗑️", key=f"delete_{item['id']}", help="删除"):
                        st.session_state.delete_item_id = item['id']
                        st.rerun()
                
                st.markdown("---")

if "edit_item_id" in st.session_state:
    item_to_edit = next((item for item in load_data() if item["id"] == st.session_state.edit_item_id), None)
    
    if item_to_edit:
        @st.dialog("编辑商品")
        def edit_item_dialog(item):
            name = st.text_input("商品名称", value=item["name"])
            categories = load_categories()
            category_index = categories.index(item["category"]) if item["category"] in categories else 0
            category = st.selectbox("分类", categories, index=category_index)
            
            col1, col2 = st.columns(2)
            quantity = col1.number_input("数量", min_value=1, value=item["quantity"])
            unit = col2.text_input("单位", value=item["unit"])
            
            budget_price = st.number_input("预算价格", min_value=0.0, value=item["budget_price"], step=0.01)
            priority_index = ["必须", "想买", "可选"].index(item["priority"]) if item["priority"] in ["必须", "想买", "可选"] else 0
            priority = st.selectbox("优先级", ["必须", "想买", "可选"], index=priority_index)
            
            platforms = ["淘宝", "天猫", "京东", "拼多多", "其他"]
            platform_index = platforms.index(item["platform"]) if item["platform"] in platforms else 0
            platform = st.selectbox("购买平台", platforms, index=platform_index)
            
            notes = st.text_area("备注", value=item.get("notes", ""))
            
            if item["status"] == "已购买":
                actual_price = st.number_input("实际价格", min_value=0.0, value=item.get("actual_price", 0), step=0.01)
            else:
                actual_price = item.get("actual_price", 0)
            
            col1, col2 = st.columns(2)
            if col1.button("保存修改", use_container_width=True):
                update_item(
                    item["id"],
                    name=name,
                    category=category,
                    quantity=quantity,
                    unit=unit,
                    budget_price=budget_price,
                    priority=priority,
                    platform=platform,
                    notes=notes,
                    actual_price=actual_price
                )
                st.success("修改成功!")
                del st.session_state.edit_item_id
                st.rerun()
            
            if col2.button("取消", use_container_width=True):
                del st.session_state.edit_item_id
                st.rerun()
        
        edit_item_dialog(item_to_edit)

if "delete_item_id" in st.session_state:
    item_to_delete = next((item for item in load_data() if item["id"] == st.session_state.delete_item_id), None)
    
    if item_to_delete:
        @st.dialog("确认删除")
        def delete_item_dialog(item):
            st.warning(f"⚠️ 定要删除商品 **{item['name']}** 吗? 删除后不可恢复!")
            
            col1, col2 = st.columns(2)
            if col1.button("确认删除", use_container_width=True):
                delete_item(item["id"])
                st.success("删除成功!")
                del st.session_state.delete_item_id
                st.rerun()
            
            if col2.button("取消", use_container_width=True):
                del st.session_state.delete_item_id
                st.rerun()
        
        delete_item_dialog(item_to_delete)

with tab2:
    st.markdown("### 📊 统计分析")
    
    all_items = load_data()
    
    if not all_items:
        st.info("暂无数据,请先添加商品")
    else:
        st.markdown("#### 💰 预算使用情况")
        budget_summary = calculate_budget_summary(all_items)
        
        col1, col2, col3 = st.columns(3)
        col1.metric("总预算", f"¥{budget_summary['total_budget']:.2f}")
        col2.metric("已花费", f"¥{budget_summary['spent']:.2f}")
        col3.metric("剩余预算", f"¥{budget_summary['remaining']:.2f}")
        
        if budget_summary['total_budget'] > 0:
            progress = budget_summary['spent'] / budget_summary['total_budget']
            st.progress(min(progress, 1.0))
            st.caption(f"预算使用率: {progress*100:.1f}%")
        
        st.markdown("---")
        
        st.markdown("#### ✅ 购买完成率")
        completion_rate = calculate_completion_rate(all_items)
        st.metric("完成率", f"{completion_rate:.1f}%")
        
        st.markdown("---")
        
        st.markdown("#### 📂 按分类统计")
        category_stats = calculate_category_stats(all_items)
        
        for category, stats in category_stats.items():
            with st.expander(f"{category} ({stats['count']}件)"):
                col1, col2, col3 = st.columns(3)
                col1.metric("预算", f"¥{stats['budget']:.2f}")
                col2.metric("已花费", f"¥{stats['spent']:.2f}")
                col3.metric("商品数", stats['count'])
        
        st.markdown("---")
        
        st.markdown("#### 🎯 按优先级统计")
        priority_stats = calculate_priority_stats(all_items)
        
        for priority, stats in priority_stats.items():
            priority_emoji = {"必须": "🔴", "想买": "🟡", "可选": "⚪"}
            with st.expander(f"{priority_emoji.get(priority, '')} {priority} ({stats['count']}件)"):
                col1, col2, col3 = st.columns(3)
                col1.metric("预算", f"¥{stats['budget']:.2f}")
                col2.metric("已花费", f"¥{stats['spent']:.2f}")
                col3.metric("商品数", stats['count'])

with tab3:
    st.markdown("### 📥 导出购物清单")
    
    all_items = load_data()
    
    if not all_items:
        st.info("暂无数据,无法导出")
    else:
        st.write(f"当前清单共 **{len(all_items)}** 件商品")
        
        if filtered_items and len(filtered_items) != len(all_items):
            st.write(f"筛选后共 **{len(filtered_items)}** 件商品")
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 导出全部商品")
            excel_data = export_to_excel(all_items)
            if excel_data:
                st.download_button(
                    label="📥 下载全部清单",
                    data=excel_data,
                    file_name=f"购物清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        with col2:
            st.markdown("#### 导出筛选后商品")
            if filtered_items and len(filtered_items) != len(all_items):
                excel_data_filtered = export_to_excel(filtered_items)
                if excel_data_filtered:
                    st.download_button(
                        label="📥 下载筛选清单",
                        data=excel_data_filtered,
                        file_name=f"购物清单_筛选_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.info("未应用筛选条件")